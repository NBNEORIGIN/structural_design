import openpyxl

# Load the CORRECTED workbook
wb = openpyxl.load_workbook('Wind_Loading_Validation_CORRECTED_20251203_1804.xlsx', data_only=True)

# Check Validation sheet
ws = wb['Validation']

print("\nVALIDATION SHEET RESULTS:")
print("="*80)
print(f"{'Parameter':<20} {'Calculated':>12} {'Expected':>12} {'Diff':>10} {'%Diff':>8} {'Status':>10}")
print("-"*80)

for row in range(5, 14):
    param = ws[f'B{row}'].value
    calc = ws[f'C{row}'].value
    expected = ws[f'D{row}'].value
    diff = ws[f'E{row}'].value
    pct = ws[f'F{row}'].value
    status = ws[f'G{row}'].value
    
    # Handle None values and errors
    try:
        if calc is None or expected is None or isinstance(calc, str) or isinstance(expected, str):
            print(f"{param:<20} {str(calc):>12} {str(expected):>12} {str(diff):>10} {str(pct):>8} {status if status else 'N/A':>10}")
        else:
            calc_val = float(calc) if calc else 0
            exp_val = float(expected) if expected else 0
            diff_val = float(diff) if diff else 0
            pct_val = float(pct) if pct and pct != '#DIV/0!' else 0
            print(f"{param:<20} {calc_val:>12.3f} {exp_val:>12.3f} {diff_val:>10.3f} {pct_val:>7.1f}% {status if status else 'N/A':>10}")
    except Exception as e:
        print(f"{param:<20} ERROR: {e}")

print("="*80)
print(f"\nOVERALL STATUS: {ws['C15'].value}")

# Now check with formulas
print("\n\nFORMULA CHECK:")
print("="*80)
wb_formula = openpyxl.load_workbook('Wind_Loading_Validation_CORRECTED_20251203_1804.xlsx')
ws_formula = wb_formula['Validation']

for row in range(5, 14):
    param = ws_formula[f'B{row}'].value
    calc_formula = ws_formula[f'C{row}'].value
    expected_formula = ws_formula[f'D{row}'].value
    status_formula = ws_formula[f'G{row}'].value
    
    print(f"\n{param}:")
    print(f"  Calc formula: {calc_formula}")
    print(f"  Expected formula: {expected_formula}")
    print(f"  Status formula: {status_formula}")

# Check Calculations sheet references
print("\n\nCALCULATIONS SHEET KEY VALUES:")
print("="*80)
ws_calc = wb['Calculations']

key_cells = {
    'C7': 'v_map calculated',
    'C8': 'v_map expected',
    'C15': 'c_alt calculated',
    'C16': 'c_alt expected',
    'C21': 'c_dir calculated',
    'C22': 'c_dir expected',
    'C36': 'c_e*c_e,T calculated',
    'C37': 'c_e*c_e,T expected',
    'C50': 'q_p calculated',
    'C51': 'q_p expected',
    'C56': 'c_s calculated',
    'C57': 'c_s expected',
    'C62': 'c_d calculated',
    'C63': 'c_d expected',
    'C68': 'c_f calculated',
    'C69': 'c_f expected',
    'C80': 'F_w calculated',
    'C81': 'F_w expected',
}

for cell, desc in key_cells.items():
    value = ws_calc[cell].value
    print(f"{cell} ({desc}): {value}")
