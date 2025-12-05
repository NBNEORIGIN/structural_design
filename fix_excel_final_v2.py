import openpyxl
from openpyxl.styles import PatternFill

# Load the workbook
wb = openpyxl.load_workbook('Wind_Loading_Validation_CORRECTED_20251203_1804.xlsx')

print("Fixing Excel workbook...")

# Fix Inputs sheet
ws_inputs = wb['Inputs']
calc_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

# Check for merged cells and unmerge if necessary
print("\nChecking merged cells in Inputs sheet...")
merged_ranges = list(ws_inputs.merged_cells.ranges)
for merged_range in merged_ranges:
    if 'C15' in str(merged_range) or 'C16' in str(merged_range) or 'C17' in str(merged_range):
        print(f"Unmerging: {merged_range}")
        ws_inputs.unmerge_cells(str(merged_range))

# Now set the formulas
ws_inputs['C15'].value = '=C5*C6'
ws_inputs['C15'].fill = calc_fill
ws_inputs['C15'].number_format = '0.0'
print("✓ Fixed C15 (A_ref)")

ws_inputs['C16'].value = '=C6/C7'
ws_inputs['C16'].fill = calc_fill
ws_inputs['C16'].number_format = '0.000'
print("✓ Fixed C16 (h/d)")

ws_inputs['C17'].value = '=C6/C5'
ws_inputs['C17'].fill = calc_fill
ws_inputs['C17'].number_format = '0.000'
print("✓ Fixed C17 (h/b)")

# Fix Calculations sheet
ws_calc = wb['Calculations']

# Fix c_f: Row 62 should reference Inputs!C16 (h/d)
ws_calc['C62'].value = '=Inputs!C16'
ws_calc['C62'].number_format = '0.000'
print("✓ Fixed C62 (h/d reference)")

# Fix F_w: Row 74 formula references wrong rows
ws_calc['C74'].value = '=C68*C69*C70*C71*C72/1000'
ws_calc['C74'].number_format = '0.0'
print("✓ Fixed C74 (F_w formula)")

# Save
filename = 'Wind_Loading_Validation_FINAL_20251203.xlsx'
wb.save(filename)
print(f"\n✅ All fixes applied and saved to: {filename}")
print("\nOpen in Excel - all validations should now PASS!")
