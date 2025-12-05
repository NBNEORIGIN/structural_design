import openpyxl

wb = openpyxl.load_workbook('Wind_Loading_Validation_CORRECTED_20251203_1804.xlsx')
ws = wb['Inputs']

print('Inputs sheet rows 14-20:')
for r in range(14, 21):
    print(f'Row {r}: B={ws[f"B{r}"].value} | C={ws[f"C{r}"].value} | D={ws[f"D{r}"].value} | E={ws[f"E{r}"].value}')
    print(f'         C{r} type: {type(ws[f"C{r}"]).__name__}')
