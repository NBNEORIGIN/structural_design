import openpyxl
from openpyxl.styles import PatternFill

# Load the workbook
wb = openpyxl.load_workbook('Wind_Loading_Validation_CORRECTED_20251203_1804.xlsx')

print("Fixing Excel workbook...")

# Fix Inputs sheet
ws_inputs = wb['Inputs']
calc_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

# Row 15: A_ref formula
ws_inputs['C15'].value = '=C5*C6'
ws_inputs['C15'].fill = calc_fill
ws_inputs['C15'].number_format = '0.0'
print("✓ Fixed C15 (A_ref)")

# Row 16: h/d formula  
ws_inputs['C16'].value = '=C6/C7'
ws_inputs['C16'].fill = calc_fill
ws_inputs['C16'].number_format = '0.000'
print("✓ Fixed C16 (h/d)")

# Row 17: h/b formula
ws_inputs['C17'].value = '=C6/C5'
ws_inputs['C17'].fill = calc_fill
ws_inputs['C17'].number_format = '0.000'
print("✓ Fixed C17 (h/b)")

# Fix Calculations sheet
ws_calc = wb['Calculations']

# Fix c_f: Row 62 should reference Inputs!C16 (h/d)
ws_calc['C62'].value = '=Inputs!C16'
ws_calc['C62'].number_format = '0.000'
print("✓ Fixed C62 (h/d reference)")

# Fix F_w: Row 74 formula references wrong rows
# Should be: q_p(C68) * c_s(C69) * c_d(C70) * c_f(C71) * A_ref(C72)
ws_calc['C74'].value = '=C68*C69*C70*C71*C72/1000'
ws_calc['C74'].number_format = '0.0'
print("✓ Fixed C74 (F_w formula)")

# Save
filename = 'Wind_Loading_Validation_FINAL_20251203.xlsx'
wb.save(filename)
print(f"\n✅ All fixes applied and saved to: {filename}")
print("\nOpen in Excel and all validations should now PASS!")
